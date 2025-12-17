# datacure_cleaning_app.py
# -----------------------------------------------------------------------------
# Datacure – prototype Streamlit
# - Import (CSV / Excel / JSON / Stata)
# - Aperçu du dataset (source de vérité: st.session_state["df"])
# - 2 modes : méthodologique (guidé) vs libre
# - Nettoyage simple (méthodologique) : standardisation texte + NA par variable (validation bouton)
# - Q&A dataset (principalement déterministe)
# - Option IA : générer code pandas de nettoyage (mode libre)
# - Export CSV + codebook Excel (remplace le PDF)
# -----------------------------------------------------------------------------

from __future__ import annotations

import io
import os
import unicodedata
from datetime import datetime
from typing import Optional, Tuple

import pandas as pd
import streamlit as st
from openai import OpenAI

# === Excel codebook (optionnel) ===
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


# === Streamlit ===
st.set_page_config(page_title="Datacure Prototype", layout="wide")
st.title("Datacure - Assistant de nettoyage de données (v0)")


# === OpenAI key ===
api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    try:
        api_key = st.secrets.get("OPENAI_API_KEY")
    except Exception:
        api_key = None

client: Optional[OpenAI] = OpenAI(api_key=api_key) if api_key else None


# === Upload ===
uploaded_file = st.file_uploader(
    "Charge un fichier de données",
    type=["csv", "xlsx", "xls", "json", "dta"],
)


def _remove_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _standardize_text_value(
    x,
    remove_accents: bool,
    acronyms: set[str],
    style: str,
    remove_double_spaces: bool = True,
) -> object:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return x
    if not isinstance(x, str):
        return x

    s = x.strip()
    if not s:
        return s

    if remove_accents:
        s = _remove_accents(s)

    if remove_double_spaces:
        s = " ".join(s.split())

    if style == "Commencer par une majuscule":
        s = s.lower().capitalize()
    elif style == "Tout en MAJUSCULES":
        s = s.upper()
    elif style == "Tout en minuscules":
        s = s.lower()
    else:
        s = s.lower().title()

    if acronyms:
        tokens = s.split(" ")
        tokens = [t.upper() if t.upper() in acronyms else t for t in tokens]
        s = " ".join(tokens)

    return s


def _text_columns(dataframe: pd.DataFrame) -> list[str]:
    return [
        c
        for c in dataframe.columns
        if dataframe[c].dtype == "object" or str(dataframe[c].dtype) == "string"
    ]


def load_data(file) -> Tuple[pd.DataFrame, str]:
    filename = (getattr(file, "name", "") or "").lower().strip()

    if filename.endswith(".csv"):
        return pd.read_csv(file), "csv"

    if filename.endswith((".xls", ".xlsx")):
        xls = pd.ExcelFile(file)  # openpyxl est utilisé par pandas si dispo
        sheet = st.selectbox("Choisis une feuille Excel", xls.sheet_names, index=0)
        return pd.read_excel(xls, sheet_name=sheet), "excel"

    if filename.endswith(".json"):
        try:
            return pd.read_json(file), "json"
        except ValueError:
            try:
                file.seek(0)
            except Exception:
                pass
            return pd.read_json(file, lines=True), "json"

    if filename.endswith(".dta"):
        return pd.read_stata(file), "stata"

    raise ValueError("Format non supporté. Utilise CSV, Excel, JSON ou Stata (.dta).")


def _reset_to_uploaded_file() -> None:
    uploaded_file.seek(0)
    df0, ft0 = load_data(uploaded_file)
    st.session_state["df"] = df0
    st.session_state["file_type"] = ft0


def _ensure_state() -> None:
    st.session_state.setdefault("cleaning_log", [])
    st.session_state.setdefault("missing_decisions", {})
    st.session_state.setdefault("missing_processed", set())
    st.session_state.setdefault("type_overrides", {})  # ✅ AJOUT


def _log_event(**kwargs) -> None:
    st.session_state["cleaning_log"].append(
        {"ts": datetime.now().isoformat(timespec="seconds"), **kwargs}
    )


def _infer_semantic_type(s: pd.Series) -> str:
    """Heuristique légère : bool/catégorielle ; num avec peu de modalités => catégorielle sinon continue."""
    if s.dtype == "bool":
        return "Catégorielle"

    if pd.api.types.is_numeric_dtype(s):
        nunique = int(s.nunique(dropna=True))
        n = int(s.notna().sum())
        if n == 0:
            return "Continue"
        if nunique <= 12:
            return "Catégorielle"
        if (nunique / max(n, 1)) <= 0.05:
            return "Catégorielle"
        return "Continue"

    return "Catégorielle"


def _detect_special_codes(s: pd.Series) -> list[tuple[str, int]]:
    """Retourne jusqu'à 6 codes spéciaux détectés sous forme [(val, count), ...]."""

    candidates = {
        -9,
        -8,
        -7,
        -6,
        -5,
        -4,
        -3,
        -2,
        -1,
        88,
        99,
        777,
        888,
        999,
        9999,
    }

    out: list[tuple[str, int]] = []

    try:
        if pd.api.types.is_numeric_dtype(s):
            for v in candidates:
                cnt = int((s == v).sum())
                if cnt > 0:
                    out.append((str(v), cnt))
        else:
            text_candidates = {
                "NA",
                "N/A",
                "NULL",
                "NONE",
                "MISSING",
                "UNK",
                "UNKNOWN",
                "-9",
                "-4",
                "99",
                "888",
                "999",
            }
            s2 = s.astype("string")
            for v in text_candidates:
                cnt = int((s2 == v).sum())
                if cnt > 0:
                    out.append((v, cnt))
    except Exception:
        return []

    out.sort(key=lambda x: x[1], reverse=True)
    return out[:6]


def _truncate(s: str, max_len: int) -> str:
    s = s or ""
    s = " ".join(s.split())
    return (s[: max_len - 1] + "…") if len(s) > max_len else s

def _col_profile(df_: pd.DataFrame, col: str) -> dict:
    s = df_[col]
    n = len(df_)
    na = int(s.isna().sum())
    nunique = int(s.nunique(dropna=True))

    # 1) Détection automatique
    try:
        semantic = _infer_semantic_type(s, col)  # si ta fonction accepte (s, col)
    except TypeError:
        semantic = _infer_semantic_type(s)       # fallback si elle n'accepte que (s)

    # 2) Override manuel (menu déroulant)
    override = st.session_state.get("type_overrides", {}).get(col)
    if override in ("Catégorielle", "Continue"):
        semantic = override

    profile = {
        "column": col,
        "dtype": str(s.dtype),
        "type": semantic,
        "missing": na,
        "missing_pct": round((na / n * 100) if n else 0.0, 2),
        "unique": nunique,
    }

    try:
        vc = s.value_counts(dropna=True)
        top_vals = [str(v) for v in vc.head(5).index.tolist()]
    except Exception:
        top_vals = []

    profile["examples"] = _truncate(", ".join(top_vals), 70)

    specials = _detect_special_codes(s)
    profile["special_codes"] = (
        _truncate(", ".join([f"{v} (n={c})" for v, c in specials]), 60) if specials else "—"
    )

    if semantic == "Continue" and pd.api.types.is_numeric_dtype(s) and s.notna().any():
        profile["min"] = float(s.min(skipna=True))
        profile["max"] = float(s.max(skipna=True))
        profile["mean"] = float(s.mean(skipna=True))
        profile["median"] = float(s.median(skipna=True))
    else:
        profile["min"] = profile["max"] = profile["mean"] = profile["median"] = None

    return profile


def _make_codebook_excel(df_: pd.DataFrame, dataset_name: str) -> bytes:
    """Génère un codebook Excel lisible, colonnes auto-ajustées, et codes spéciaux détectés."""
    if not _HAS_OPENPYXL:
        raise ModuleNotFoundError("openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Codebook"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Titre (sur 11 colonnes)
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = dataset_name
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    # Infos générales
    ws["A3"] = f"Nombre de lignes: {len(df_):,}"
    ws["A4"] = f"Nombre de colonnes: {len(df_.columns):,}"

    headers = [
        "#",
        "Variable",
        "Type",
        "Non-null",
        "Null (%)",
        "Unique",
        "Exemples (top 5)",
        "Codes spéciaux détectés",
        "Min",
        "Max",
        "Moy./Med.",
    ]

    start_row = 6

    # En-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Figer l'entête + filtre
    ws.freeze_panes = ws[f"A{start_row+1}"]
    ws.auto_filter.ref = f"A{start_row}:K{start_row}"

    # Lignes
    for i, col in enumerate(df_.columns, start=1):
        p = _col_profile(df_, str(col))
        row_idx = start_row + i

        non_null = int(df_[col].notna().sum())
        null_pct = float((df_[col].isna().sum() / len(df_) * 100) if len(df_) else 0.0)

        # Moy./Med. uniquement si continue
        mm = "—"
        if p["mean"] is not None and p["median"] is not None:
            mm = f"mean={p['mean']:.4g}; med={p['median']:.4g}"

        row_data = [
            i,
            p["column"],
            p["type"],
            non_null,
            f"{null_pct:.2f}%",
            p["unique"],
            p["examples"],
            p["special_codes"],
            "" if p["min"] is None else f"{p['min']:.4g}",
            "" if p["max"] is None else f"{p['max']:.4g}",
            mm,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Mise en évidence si codes spéciaux détectés
            if col_idx == 8 and isinstance(value, str) and value not in ("", "—"):
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Auto-ajustement des largeurs (sans se faire biaiser par le titre et A3/A4)
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in range(start_row, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_length = max(max_length, len(str(v)))

        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

    # Sauvegarde buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# === Pas de fichier ===
if not uploaded_file:
    st.info("📂 Veuillez charger un fichier (CSV, Excel, JSON ou Stata) pour commencer.")
    st.stop()


# === Lecture + session state ===
try:
    df_in, file_type_in = load_data(uploaded_file)
except Exception as e:
    st.error(f"Erreur de lecture du fichier : {e}")
    st.stop()

uploaded_name = getattr(uploaded_file, "name", None)
if st.session_state.get("uploaded_name") != uploaded_name:
    st.session_state["df"] = df_in
    st.session_state["file_type"] = file_type_in
    st.session_state["uploaded_name"] = uploaded_name
    st.session_state.pop("generated_code", None)
    st.session_state.pop("missing_decisions", None)
    st.session_state.pop("missing_processed", None)
    st.session_state.pop("cleaning_log", None)

_ensure_state()

df = st.session_state.get("df")
file_type = st.session_state.get("file_type", "inconnu")


# === Aperçu ===
st.subheader("Aperçu du fichier")
st.caption(f"📄 Format détecté : {file_type}")
st.dataframe(df.head())


# === Mode ===
mode = st.radio(
    "Mode de travail",
    ["🧭 Nettoyage méthodologique", "🧪 Mode libre"],
    horizontal=True,
    key="mode",
)

# --- Typage manuel (override) + value labels : visible dans les 2 modes ---
with st.expander("🏷️ Labellisation des variables (Catégorielle vs Continue)", expanded=False):
    st.caption("Force le type pour les variables numériques catégorielles et ajoute des labels aux modalités.")

    # candidats (numériques uniquement) pour le typage
    candidates = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    if not candidates:
        st.info("Aucune colonne numérique détectée.")
    else:
        # ✅ on travaille en strings partout pour éviter les mismatches
        options = [str(c) for c in candidates]

        cols = st.multiselect(
            "Sélectionne les colonnes à typer",
            options=options,
            default=[],
            key="type_cols_select",
        )

        # ✅ Batch apply
        batch_value = st.selectbox(
            "Appliquer à toutes les colonnes sélectionnées",
            ["—", "Auto (détection)", "Catégorielle", "Continue"],
            index=0,
            key="type_batch_value",
        )

        if st.button("⚡ Appliquer à toutes", use_container_width=True, key="type_batch_apply"):
            if batch_value != "—":
                for col in cols:
                    # met aussi à jour les selectbox individuels
                    st.session_state[f"type_override__{col}"] = batch_value

                    # met à jour le stockage
                    if batch_value == "Auto (détection)":
                        st.session_state["type_overrides"].pop(col, None)
                    else:
                        st.session_state["type_overrides"][col] = batch_value

                st.success("✅ Type appliqué à toutes les colonnes sélectionnées.")
                st.rerun()

        # Réglage fin (optionnel) colonne par colonne
        for col in cols:
            current = st.session_state.get("type_overrides", {}).get(col, "Auto (détection)")
            st.selectbox(
                f"{col}",
                ["Auto (détection)", "Catégorielle", "Continue"],
                index=["Auto (détection)", "Catégorielle", "Continue"].index(current),
                key=f"type_override__{col}",
            )

        a, b = st.columns(2)
        with a:
            if st.button("✅ Appliquer le typage", use_container_width=True, key="type_apply"):
                for col in cols:
                    v = st.session_state.get(f"type_override__{col}", "Auto (détection)")
                    if v == "Auto (détection)":
                        st.session_state["type_overrides"].pop(col, None)
                    else:
                        st.session_state["type_overrides"][col] = v
                st.success("Typage enregistré.")
                st.rerun()

        with b:
            if st.button("🧹 Effacer tous les overrides", use_container_width=True, key="type_clear"):
                st.session_state["type_overrides"] = {}
                st.success("Overrides supprimés.")
                st.rerun()

    # =========================
    # Value labels (modalités)
    # =========================
    st.divider()
    st.markdown("### 🏷️ Nommer les valeurs (modalités) d’une variable catégorielle")

    label_col = st.selectbox(
        "Variable à labelliser",
        options=[str(c) for c in df.columns],
        key="vl_col",
    )

    s = df[label_col]

    # Limite pour éviter une table énorme
    n_unique = int(s.nunique(dropna=True))
    max_modalities = 200
    if n_unique > max_modalities:
        st.warning(f"⚠️ {label_col} a {n_unique} valeurs uniques. Affichage limité aux {max_modalities} plus fréquentes.")

    # Modalités (top par fréquence)
    vals = (
        s.dropna()
        .astype("string")
        .value_counts()
        .head(max_modalities)
        .index
        .tolist()
    )

    existing = st.session_state.get("value_labels", {}).get(label_col, {})
    vl_df = pd.DataFrame(
        {
            "value": vals,
            "label": [existing.get(v, "") for v in vals],
        }
    )

    edited = st.data_editor(
        vl_df,
        use_container_width=True,
        num_rows="fixed",
        key="vl_editor",
    )

    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("✅ Enregistrer les labels", use_container_width=True, key="vl_save"):
            mapping = {}
            for _, r in edited.iterrows():
                v = str(r["value"])
                lab = str(r["label"]).strip()
                if lab:
                    mapping[v] = lab

            st.session_state.setdefault("value_labels", {})
            st.session_state["value_labels"][label_col] = mapping
            st.success(f"Labels enregistrés pour {label_col} ({len(mapping)} valeur(s)).")

    with c2:
        if st.button("🧾 Appliquer au dataset", use_container_width=True, key="vl_apply"):
            mapping = st.session_state.get("value_labels", {}).get(label_col, {})
            if not mapping:
                st.warning("Aucun label enregistré pour cette variable.")
            else:
                df2 = st.session_state["df"].copy()
                new_col = f"{label_col}_label"
                base = df2[label_col].astype("string")
                df2[new_col] = base.map(mapping).fillna(base)
                st.session_state["df"] = df2
                st.success(f"Colonne créée : {new_col}")
                st.rerun()

    with c3:
        if st.button("🗑️ Supprimer ces labels", use_container_width=True, key="vl_clear"):
            st.session_state.get("value_labels", {}).pop(label_col, None)
            st.success(f"Labels supprimés pour {label_col}.")
            st.rerun()

# ================================
# 🧭 MODE MÉTHODOLOGIQUE (simple)
# ================================
if mode == "🧭 Nettoyage méthodologique":
    with st.expander("🧭 Nettoyage simple", expanded=True):
        st.caption("Nettoyage sûr (bouton par bouton), sans interprétation statistique")

        # --- 1) Standardiser le texte ---
        st.markdown("### 1) 🧹 Standardiser le texte")
        cols_text = _text_columns(df)

        style = st.selectbox(
            "Style",
            [
                "Majuscule à chaque mot",
                "Commencer par une majuscule",
                "Tout en MAJUSCULES",
                "Tout en minuscules",
            ],
            index=0,
            key="m_std_style",
        )

        remove_acc = st.checkbox("Supprimer les accents", value=True, key="m_std_acc")
        remove_double_spaces = st.checkbox(
            "Supprimer les doubles espaces", value=True, key="m_std_spaces"
        )

        acronyms_raw = st.text_input(
            "Acronymes à garder en MAJ (séparés par des virgules)",
            value="",
            key="m_std_acronyms",
        )
        acronyms = {a.strip().upper() for a in acronyms_raw.split(",") if a.strip()}

        scope = st.radio(
            "Appliquer sur",
            ["Tout le tableau", "Une colonne", "Une ligne"],
            horizontal=True,
            key="m_std_scope",
        )

        selected_col: Optional[str] = None
        if scope == "Une colonne":
            selected_col = st.selectbox(
                "Colonne",
                cols_text if cols_text else ["(aucune colonne texte)"],
                key="m_std_col",
            )
            if not cols_text:
                selected_col = None

        selected_row: Optional[int] = None
        if scope == "Une ligne":
            selected_row = int(
                st.number_input(
                    "Index de ligne (0 = première ligne)",
                    min_value=0,
                    max_value=max(0, len(df) - 1),
                    value=0,
                    step=1,
                    key="m_std_row",
                )
            )

        c1, c2 = st.columns(2)
        with c1:
            if st.button(
                "✨ Appliquer la standardisation",
                use_container_width=True,
                key="m_std_apply",
            ):
                if not cols_text:
                    st.warning("Aucune colonne texte détectée.")
                else:
                    df2 = df.copy()
                    if scope == "Tout le tableau":
                        for c in cols_text:
                            df2[c] = df2[c].apply(
                                lambda v: _standardize_text_value(
                                    v,
                                    remove_acc,
                                    acronyms,
                                    style,
                                    remove_double_spaces,
                                )
                            )
                        _log_event(
                            step="standardiser_texte",
                            scope="tableau",
                            columns=cols_text,
                            options={
                                "style": style,
                                "remove_acc": remove_acc,
                                "remove_double_spaces": remove_double_spaces,
                                "acronyms": sorted(list(acronyms)),
                            },
                        )
                    elif scope == "Une colonne" and selected_col:
                        df2[selected_col] = df2[selected_col].apply(
                            lambda v: _standardize_text_value(
                                v,
                                remove_acc,
                                acronyms,
                                style,
                                remove_double_spaces,
                            )
                        )
                        _log_event(
                            step="standardiser_texte",
                            scope="colonne",
                            column=selected_col,
                            options={
                                "style": style,
                                "remove_acc": remove_acc,
                                "remove_double_spaces": remove_double_spaces,
                                "acronyms": sorted(list(acronyms)),
                            },
                        )
                    elif scope == "Une ligne" and selected_row is not None:
                        r = selected_row
                        for c in cols_text:
                            df2.at[df2.index[r], c] = _standardize_text_value(
                                df2.at[df2.index[r], c],
                                remove_acc,
                                acronyms,
                                style,
                                remove_double_spaces,
                            )
                        _log_event(
                            step="standardiser_texte",
                            scope="ligne",
                            row_index=int(r),
                            columns=cols_text,
                            options={
                                "style": style,
                                "remove_acc": remove_acc,
                                "remove_double_spaces": remove_double_spaces,
                                "acronyms": sorted(list(acronyms)),
                            },
                        )

                    st.session_state["df"] = df2
                    st.success("✅ Standardisation appliquée")
                    st.rerun()

        with c2:
            if st.button("↩️ Annuler les changements", use_container_width=True, key="m_reset"):
                _reset_to_uploaded_file()
                st.session_state["missing_decisions"] = {}
                st.session_state["missing_processed"] = set()
                st.session_state["cleaning_log"] = []
                st.success("✅ Réinitialisé")
                st.rerun()

        st.divider()

        # --- 2) Valeurs manquantes (par variable) ---
        st.markdown("### 2) Valeurs manquantes (par variable)")

        n_rows = len(df)
        miss = df.isna().sum().astype(int)
        miss_pct = (miss / n_rows * 100).round(2) if n_rows else (miss * 0.0)

        miss_tbl = pd.DataFrame(
            {
                "Variable": miss.index.astype(str),
                "Type": df.dtypes.astype(str).values,
                "NA (n)": miss.values,
                "NA (%)": miss_pct.values,
                "État": [
                    "✅ validé" if str(c) in st.session_state["missing_processed"] else "⏳ à décider"
                    for c in miss.index.astype(str)
                ],
            }
        ).sort_values(["NA (n)", "Variable"], ascending=[False, True])

        def _style_missing(row):
            try:
                p = float(row["NA (%)"])
            except Exception:
                p = 0.0
            return ["font-weight:600"] * len(row) if p >= 20 else [""] * len(row)

        st.dataframe(miss_tbl.style.apply(_style_missing, axis=1), use_container_width=True)
        st.caption(
            "Astuce: les décisions s'appliquent immédiatement au dataset (suppression de lignes), variable par variable."
        )

        cols_with_na = [c for c in df.columns if int(df[c].isna().sum()) > 0]
        if not cols_with_na:
            st.success("Aucune valeur manquante détectée.")
        else:
            st.markdown("#### Décider et valider (bouton par variable)")

            for col in cols_with_na:
                col_name = str(col)
                na_n = int(df[col].isna().sum())
                na_p = float((na_n / len(df) * 100) if len(df) else 0.0)

                is_done = col_name in st.session_state["missing_processed"]

                with st.container():
                    left, right = st.columns([3, 2])

                    with left:
                        st.write(f"**{col_name}** — NA: {na_n:,} ({na_p:.2f}%)")

                    with right:
                        decision_key = f"miss_dec__{col_name}"
                        btn_key = f"miss_apply__{col_name}"

                        default_dec = st.session_state["missing_decisions"].get(
                            col_name, "Ne rien faire (garder les NA)"
                        )

                        decision = st.radio(
                            "Décision",
                            [
                                "Ne rien faire (garder les NA)",
                                f"Exclure les lignes où '{col_name}' est manquant",
                                f"Marquer '{col_name}' comme analyse en cas complets",
                            ],
                            index=[
                                "Ne rien faire (garder les NA)",
                                f"Exclure les lignes où '{col_name}' est manquant",
                                f"Marquer '{col_name}' comme analyse en cas complets",
                            ].index(default_dec),
                            key=decision_key,
                            disabled=is_done,
                            label_visibility="collapsed",
                        )

                        if st.button(
                            "✅ Valider",
                            use_container_width=True,
                            key=btn_key,
                            disabled=is_done,
                        ):
                            st.session_state["missing_decisions"][col_name] = decision

                            before = len(df)
                            rows_removed = 0

                            if decision.startswith("Exclure les lignes"):
                                df2 = df.dropna(subset=[col])
                                rows_removed = before - len(df2)
                                st.session_state["df"] = df2
                                df = df2

                            st.session_state["missing_processed"].add(col_name)

                            _log_event(
                                step="valeurs_manquantes",
                                variable=col_name,
                                missing_n=int(na_n),
                                missing_pct=float(round(na_p, 2)),
                                decision=decision,
                                rows_removed=int(rows_removed),
                                rows_before=int(before),
                                rows_after=int(len(df)),
                            )

                            st.success(f"Décision validée pour '{col_name}'.")
                            st.rerun()

                    st.divider()


# ================================
# 🧪 MODE LIBRE
# ================================
else:
    with st.expander("💬 Poser une question sur le dataset", expanded=False):
        q = st.text_input(
            "Question",
            placeholder="Ex : Combien de lignes ? Y a-t-il des doublons d'identifiant ?",
            key="qa_q",
        )

        id_candidates = list(df.columns)
        id_col = st.selectbox("Colonne identifiant (si besoin)", id_candidates, key="qa_id")

        qa_out = st.empty()

        def _answer_question(question: str) -> None:
            qq = (question or "").strip().lower()
            if not qq:
                qa_out.info("Écris une question ci-dessus.")
                return

            if "combien" in qq and ("ligne" in qq or "rows" in qq):
                qa_out.success(f"Il y a {len(df):,} ligne(s).")
                return
            if "combien" in qq and ("colonne" in qq or "columns" in qq):
                qa_out.success(f"Il y a {df.shape[1]:,} colonne(s).")
                return
            if "dimension" in qq or "shape" in qq:
                qa_out.success(f"Dimensions : {df.shape[0]:,} lignes × {df.shape[1]:,} colonnes")
                return

            if "fi item id" in qq or "fi_item_id" in qq or "fi-item id" in qq:
                col2 = None
                if "FI Item ID" in df.columns:
                    col2 = "FI Item ID"
                else:
                    norm = {str(c).strip().lower(): c for c in df.columns}
                    col2 = norm.get("fi item id") or norm.get("fi_item_id") or norm.get("fi-item id")

                if not col2:
                    qa_out.error("Je ne trouve pas la colonne 'FI Item ID' dans ce dataset.")
                    return

                n_unique = int(df[col2].nunique(dropna=True))
                qa_out.success(f"'{col2}' contient {n_unique:,} identifiant(s) unique(s) (hors NA).")
                return

            if "doubl" in qq or "duplicate" in qq or ("identifi" in qq and "2 fois" in qq):
                if id_col not in df.columns:
                    qa_out.error("Je ne trouve pas la colonne identifiant sélectionnée.")
                    return
                s = df[id_col]
                dup_mask = s.duplicated(keep=False) & s.notna()
                n_dup_rows = int(dup_mask.sum())
                n_dup_ids = int(s[dup_mask].nunique(dropna=True))
                if n_dup_rows == 0:
                    qa_out.success(f"Aucun doublon détecté dans '{id_col}'.")
                else:
                    qa_out.warning(
                        f"Doublons détectés dans '{id_col}' : {n_dup_ids} identifiant(s) dupliqué(s), "
                        f"touchant {n_dup_rows} ligne(s)."
                    )
                    sample_ids = s[dup_mask].astype("string").dropna().unique()[:10]
                    qa_out.write("Exemples d'identifiants dupliqués :")
                    qa_out.write(list(sample_ids))
                return

            if "manquant" in qq or "missing" in qq or " na" in f" {qq} ":
                na_counts = df.isna().sum().sort_values(ascending=False)
                top = na_counts[na_counts > 0].head(20)
                if top.empty:
                    qa_out.success("Aucune valeur manquante détectée.")
                else:
                    qa_out.warning("Colonnes avec des valeurs manquantes (top 20) :")
                    qa_out.dataframe(top.to_frame(name="NA"))
                return

            if "unique" in qq or "distinct" in qq:
                if id_col in df.columns:
                    n_unique = int(df[id_col].nunique(dropna=True))
                    qa_out.success(f"'{id_col}' contient {n_unique:,} valeur(s) unique(s) (hors NA).")
                    return

            if client:
                dtypes_txt = df.dtypes.astype(str).to_dict()
                preview = df.head(20).to_dict(orient="records")
                prompt_qa = f"""Tu es un assistant d'analyse de données.
Réponds brièvement en français.
Ne fournis pas de code.

Question: {question}
Colonne_identifiant: {id_col}
dtypes: {dtypes_txt}
aperçu_20_lignes: {preview}
"""

                try:
                    resp = client.chat.completions.create(
                        model="gpt-3.5-turbo",
                        messages=[{"role": "user", "content": prompt_qa}],
                        temperature=0,
                    )
                    qa_out.info(resp.choices[0].message.content.strip())
                except Exception as e:
                    qa_out.error(f"Erreur IA : {e}")
            else:
                qa_out.info("Question non reconnue. Essaie : lignes, colonnes, doublons, manquants.")

        if st.button("🤖 Répondre", use_container_width=True, key="qa_btn"):
            _answer_question(q)

    with st.expander("⚡ Commandes rapides", expanded=False):
        cols_text = _text_columns(df)

        style = st.selectbox(
            "Standardiser (style)",
            [
                "Majuscule à chaque mot",
                "Commencer par une majuscule",
                "Tout en MAJUSCULES",
                "Tout en minuscules",
            ],
            index=0,
            key="f_std_style",
        )
        remove_acc = st.checkbox("Supprimer les accents", value=True, key="f_std_acc")
        remove_double_spaces = st.checkbox(
            "Supprimer les doubles espaces", value=True, key="f_std_spaces"
        )
        acronyms_raw = st.text_input("Acronymes (virgules)", value="", key="f_std_acronyms")
        acronyms = {a.strip().upper() for a in acronyms_raw.split(",") if a.strip()}

        scope = st.radio(
            "Appliquer sur",
            ["Tout le tableau", "Une colonne", "Une ligne"],
            horizontal=True,
            key="f_std_scope",
        )
        selected_col: Optional[str] = None
        if scope == "Une colonne":
            selected_col = st.selectbox("Colonne", cols_text if cols_text else ["(aucune)"], key="f_std_col")
            if not cols_text:
                selected_col = None

        selected_row: Optional[int] = None
        if scope == "Une ligne":
            selected_row = int(
                st.number_input(
                    "Index de ligne",
                    min_value=0,
                    max_value=max(0, len(df) - 1),
                    value=0,
                    step=1,
                    key="f_std_row",
                )
            )

        a1, a2 = st.columns(2)
        with a1:
            if st.button("✨ Standardiser", use_container_width=True, key="f_std_apply"):
                if not cols_text:
                    st.warning("Aucune colonne texte détectée.")
                else:
                    df2 = df.copy()
                    if scope == "Tout le tableau":
                        for c in cols_text:
                            df2[c] = df2[c].apply(
                                lambda v: _standardize_text_value(
                                    v,
                                    remove_acc,
                                    acronyms,
                                    style,
                                    remove_double_spaces,
                                )
                            )
                    elif scope == "Une colonne" and selected_col:
                        df2[selected_col] = df2[selected_col].apply(
                            lambda v: _standardize_text_value(
                                v,
                                remove_acc,
                                acronyms,
                                style,
                                remove_double_spaces,
                            )
                        )
                    elif scope == "Une ligne" and selected_row is not None:
                        r = selected_row
                        for c in cols_text:
                            df2.at[df2.index[r], c] = _standardize_text_value(
                                df2.at[df2.index[r], c],
                                remove_acc,
                                acronyms,
                                style,
                                remove_double_spaces,
                            )
                    st.session_state["df"] = df2
                    st.rerun()

        with a2:
            if st.button("↩️ Reset", use_container_width=True, key="f_reset"):
                _reset_to_uploaded_file()
                st.rerun()

    if not client:
        st.info("(IA désactivée) Ajoute OPENAI_API_KEY pour activer le nettoyage via API.")

    user_input = st.text_input(
        "Commande IA (optionnel)",
        placeholder="Ex : supprime les lignes où age est manquant",
        key="nl_cmd",
    )

    if user_input and client:
        prompt = f"""
Tu es un assistant Python expert en nettoyage de données avec pandas.
Voici un DataFrame nommé df.
L'utilisateur demande : \"{user_input}\".

Contraintes:
- Retourne uniquement du code Python exécutable.
- Le code doit MODIFIER le DataFrame df et laisser df comme résultat final.
- N'utilise pas d'import.
- N'accède pas au système de fichiers.
- N'utilise pas de réseau.
""".strip()

        result_container = st.empty()

        with st.spinner("🧠 Génération du code Python..."):
            try:
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0,
                )
                st.session_state["generated_code"] = response.choices[0].message.content.strip()
            except Exception as e:
                result_container.error(f"❌ Erreur lors de l'appel à l'API OpenAI : {e}")

        if "generated_code" in st.session_state:
            code = st.session_state["generated_code"]
            with st.expander("🧠 Voir le code généré", expanded=False):
                st.code(code, language="python")

            if st.button("▶️ Exécuter", key="run_ai"):
                try:
                    local_vars = {"df": df.copy()}
                    exec(code, {}, local_vars)
                    if "df" not in local_vars:
                        raise RuntimeError("Le code généré n'a pas laissé de variable 'df' en sortie.")
                    st.session_state["df"] = local_vars["df"]
                    result_container.success("✅ Nettoyage appliqué")
                    st.rerun()
                except Exception as e:
                    result_container.error(f"❌ Erreur pendant l'exécution : {e}")


# === Codebook Excel ===
with st.expander("📊 Codebook (Excel)", expanded=False):
    if not _HAS_OPENPYXL:
        st.info("Excel indisponible : installe openpyxl (pip install openpyxl).")
    else:
        default_name = st.session_state.get("uploaded_name") or "dataset"
        excel_title = st.text_input(
            "Titre du codebook",
            value=f"Codebook - {default_name}",
            key="excel_title",
        )
        if st.button("📊 Générer le codebook Excel", use_container_width=True, key="excel_btn"):
            try:
                with st.spinner("Génération du codebook Excel..."):
                    excel_bytes = _make_codebook_excel(st.session_state["df"], dataset_name=excel_title)

                st.download_button(
                    label="⬇️ Télécharger le codebook (Excel)",
                    data=excel_bytes,
                    file_name=f"{default_name}_codebook.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success("✅ Codebook Excel généré")
            except Exception as e:
                st.error(f"Erreur Excel : {e}")


# === Export CSV ===
df_final = st.session_state.get("df", df)
cleaned_csv = df_final.to_csv(index=False).encode("utf-8")

st.download_button(
    label="📥 Télécharger le fichier nettoyé (CSV)",
    data=cleaned_csv,
    file_name="fichier_nettoye.csv",
    mime="text/csv",
)


# === Mini-tests (optionnels) ===
if os.getenv("DATACURE_RUN_TESTS") == "1":
    import json

    class _FakeUpload:
        def __init__(self, name: str, payload: bytes):
            self.name = name
            self._bio = io.BytesIO(payload)

        def read(self, *args, **kwargs):
            return self._bio.read(*args, **kwargs)

        def seek(self, pos: int):
            return self._bio.seek(pos)

        def __getattr__(self, item):
            return getattr(self._bio, item)

    fake_csv = _FakeUpload("test.csv", b"a,b\n1,2\n")
    df_csv, t_csv = load_data(fake_csv)
    assert t_csv == "csv" and df_csv.shape == (1, 2)

    payload = json.dumps([{"a": 1, "b": 2}]).encode("utf-8")
    fake_json = _FakeUpload("test.json", payload)
    df_json, t_json = load_data(fake_json)
    assert t_json == "json" and df_json.shape == (1, 2)

    payload_jsonl = b"{\"a\": 1, \"b\": 2}\n{\"a\": 3, \"b\": 4}\n"
    fake_jsonl = _FakeUpload("test.json", payload_jsonl)
    df_jsonl, t_jsonl = load_data(fake_jsonl)
    assert t_jsonl == "json" and df_jsonl.shape == (2, 2)

    assert _standardize_text_value("  chu  bruxelles ", True, {"CHU"}, "Majuscule à chaque mot") == "CHU Bruxelles"
    assert _standardize_text_value("abc DEF", False, set(), "Tout en minuscules") == "abc def"
    assert _standardize_text_value("abc DEF", False, set(), "Tout en MAJUSCULES") == "ABC DEF"
    assert _standardize_text_value("abc DEF", False, set(), "Commencer par une majuscule") == "Abc def"
    assert _standardize_text_value("   ", False, set(), "Tout en MAJUSCULES") == ""

    # codes spéciaux
    s = pd.Series([1, 99, 99, None])
    assert _detect_special_codes(s)[0][0] == "99"

    st.success("✅ DATACURE_RUN_TESTS: tous les mini-tests ont réussi")
